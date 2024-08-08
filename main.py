import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

excel_path = '/content/TestData.xlsx'

# Load the Excel file using openpyxl
workbook = load_workbook(excel_path)
sheet = workbook.active

# Iterate over the rows in the sheet
for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):  # Start from row 2 to skip header
    url = row[3]  # Assuming URL is in the second column, adjust as necessary

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)

    # Save QR code to a BytesIO object
    img_stream = BytesIO()
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(img_stream, format='PNG')
    img_stream.seek(0)

    # Insert the image into the Excel file
    img = Image(img_stream)
    img.anchor = f'C{index + 2}'  # Adjust the cell reference as per your need
    sheet.add_image(img)

# Save the updated Excel file
workbook.save('/content/TestData_with_qr.xlsx')
